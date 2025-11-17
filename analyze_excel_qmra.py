#!/usr/bin/env python3
"""
Comprehensive Excel QMRA File Analysis
======================================
Extracts all formulas, data, and calculations from the Excel file.
"""

import openpyxl
from openpyxl import load_workbook
import pandas as pd
from pathlib import Path

def analyze_excel_file(excel_path):
    """Analyze Excel file and extract all formulas and data."""

    print("="*80)
    print(f"ANALYZING: {excel_path.name}")
    print("="*80)

    # Load workbook with data_only=False to get formulas
    wb = load_workbook(excel_path, data_only=False)

    print(f"\nWorksheets found: {wb.sheetnames}")
    print(f"Total sheets: {len(wb.sheetnames)}\n")

    all_formulas = {}
    all_data = {}

    for sheet_name in wb.sheetnames:
        print("\n" + "="*80)
        print(f"SHEET: {sheet_name}")
        print("="*80)

        ws = wb[sheet_name]

        # Get sheet dimensions
        print(f"Dimensions: {ws.dimensions}")
        print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")

        # Extract all formulas
        formulas = {}
        data_values = {}

        for row in ws.iter_rows():
            for cell in row:
                cell_ref = f"{cell.column_letter}{cell.row}"

                # Check if cell has a formula
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas[cell_ref] = cell.value
                elif cell.value is not None:
                    data_values[cell_ref] = cell.value

        print(f"\nFormulas found: {len(formulas)}")
        print(f"Data cells found: {len(data_values)}")

        if formulas:
            print("\n--- FORMULAS ---")
            for cell_ref, formula in sorted(formulas.items(), key=lambda x: (int(''.join(filter(str.isdigit, x[0]))), x[0])):
                print(f"  {cell_ref}: {formula}")

        # Show sample data
        print("\n--- SAMPLE DATA (first 20 cells) ---")
        for i, (cell_ref, value) in enumerate(sorted(data_values.items(), key=lambda x: (int(''.join(filter(str.isdigit, x[0]))), x[0]))[:20]):
            print(f"  {cell_ref}: {value}")

        # Store for later use
        all_formulas[sheet_name] = formulas
        all_data[sheet_name] = data_values

        # Try to read as DataFrame for structured view
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
            print(f"\n--- DATAFRAME SHAPE ---")
            print(f"  Rows: {df.shape[0]}, Columns: {df.shape[1]}")
            print("\n--- FIRST 10 ROWS ---")
            print(df.head(10).to_string())
        except Exception as e:
            print(f"\nCould not read as DataFrame: {e}")

    # Summary
    print("\n" + "="*80)
    print("SUMMARY")
    print("="*80)

    for sheet_name in wb.sheetnames:
        print(f"\n{sheet_name}:")
        print(f"  Formulas: {len(all_formulas[sheet_name])}")
        print(f"  Data cells: {len(all_data[sheet_name])}")

    # Identify key calculation patterns
    print("\n" + "="*80)
    print("KEY CALCULATIONS IDENTIFIED")
    print("="*80)

    for sheet_name, formulas in all_formulas.items():
        print(f"\n{sheet_name}:")

        # Look for specific function types
        functions_used = set()
        for formula in formulas.values():
            # Extract function names (simple pattern matching)
            import re
            funcs = re.findall(r'\b([A-Z]+)\(', formula)
            functions_used.update(funcs)

        if functions_used:
            print(f"  Functions used: {', '.join(sorted(functions_used))}")

        # Look for GAMMALN (Beta-Binomial indicator)
        gammaln_formulas = {k: v for k, v in formulas.items() if 'GAMMALN' in v}
        if gammaln_formulas:
            print(f"  GAMMALN formulas (Beta-Binomial): {len(gammaln_formulas)}")
            for cell_ref, formula in list(gammaln_formulas.items())[:3]:
                print(f"    {cell_ref}: {formula}")

    return all_formulas, all_data


def main():
    excel_path = Path("QMRA_Shellfish_191023_Nino_SUMMER.xlsx")

    if not excel_path.exists():
        print(f"ERROR: File not found: {excel_path}")
        return

    print(f"File size: {excel_path.stat().st_size / 1024:.1f} KB")
    print(f"Last modified: {pd.Timestamp.fromtimestamp(excel_path.stat().st_mtime)}")
    print()

    all_formulas, all_data = analyze_excel_file(excel_path)

    # Save detailed output to file
    output_file = Path("EXCEL_ANALYSIS_DETAILED.txt")
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("="*80 + "\n")
        f.write("COMPLETE EXCEL FORMULA EXTRACTION\n")
        f.write("="*80 + "\n\n")

        for sheet_name, formulas in all_formulas.items():
            f.write(f"\n{'='*80}\n")
            f.write(f"SHEET: {sheet_name}\n")
            f.write(f"{'='*80}\n\n")

            f.write(f"Total formulas: {len(formulas)}\n\n")

            for cell_ref, formula in sorted(formulas.items(), key=lambda x: (int(''.join(filter(str.isdigit, x[0]))), x[0])):
                f.write(f"{cell_ref}: {formula}\n")

            f.write(f"\n\nData cells:\n")
            data = all_data[sheet_name]
            for cell_ref, value in sorted(data.items(), key=lambda x: (int(''.join(filter(str.isdigit, x[0]))), x[0])):
                f.write(f"{cell_ref}: {value}\n")

    print(f"\n{'='*80}")
    print(f"Detailed analysis saved to: {output_file}")
    print(f"{'='*80}")


if __name__ == '__main__':
    main()
